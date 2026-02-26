#!/usr/bin/env python3
"""
Reflow.py — Built-in Python web server UI (no Flask)

Reflow Profiler:
- Upload BOM (.xlsx/.xlsm) -> preview table
- Component agent trigger (below BOM preview):
    - Extracts unique MPNs (MPN column only)
    - Creates scaffold results for Tp / within 5°C / TAL / ramp-up / cool-down
- Board reflow trigger (on PCB panel):
    - Placeholder: marks board-side as "ready" (no profile logic yet)
- Aggregate button:
    - Disabled until BOTH board + components activated
    - Uses slider weighting (board/components) for placeholder output

Requirements:
  py -m pip install openpyxl 
  py -m pip install pypdf

Run:
  py Reflow.py
"""

from __future__ import annotations

import io
import json
import threading
import time
import webbrowser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

import cgi  # stdlib multipart parsing

import re
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pypdf import PdfReader

HOST = "127.0.0.1"
PORT = 8000

_STATE_LOCK = threading.Lock()
_STATE = {
    "boardWeight": 80,
    "bomLabel": "No BOM loaded",
    "bom": {
        "columns": ["Item", "Stock", "Description", "Qty", "Ref", "MPN"],
        "rows": [
            {
                "Item": "1",
                "Stock": "Brynleigh0001",
                "Description": "(example)",
                "Qty": "67",
                "Ref": "FK1, FK2..",
                "MPN": "F3R76H90K",
            }
        ],
    },
    "components_ready": False,
    "board_ready": False,
    "component_reflow": [],
    "aggregation": {"status": "not_run", "summary": ""},
}

INDEX_HTML = r"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>Reflow Profiling UI</title>
  <style>
    :root{
      --brg-dark:#0b2b1a;
      --brg-main:#0f3a23;
      --brg-mid:#145b34;
      --brg-light:#e7f3ec;
      --border:rgba(11,43,26,0.15);
      --shadow:0 10px 25px rgba(11,43,26,0.05);
      --muted:#5b6b62;
      --bg:#ffffff;
    }
    *{box-sizing:border-box}
    html,body{height:100%}
    body{
      margin:0;
      font-family:ui-sans-serif,system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;
      background:var(--bg);
      color:var(--brg-dark);
    }
    .app{height:100vh;width:100vw;padding:24px}
    .frame{
      height:100%;
      max-width:1400px;
      margin:0 auto;
      display:flex;
      flex-direction:column;
      gap:16px;
    }
    .topbar{display:flex;align-items:center;gap:12px}

    .pill{
      margin-left:auto;
      border:1px solid var(--border);
      background:#fff;
      padding:6px 10px;
      border-radius:999px;
      font-size:12px;
      color:var(--brg-dark);
    }

    .btn{
      border:0;
      border-radius:12px;
      padding:10px 14px;
      font-size:14px;
      font-weight:600;
      cursor:pointer;
      box-shadow:0 2px 10px rgba(11,43,26,0.08);
    }
    .btn-primary{background:var(--brg-main);color:#fff}
    .btn-primary:hover{background:var(--brg-dark)}
    .btn-secondary{
      background:#fff;color:var(--brg-dark);
      border:1px solid rgba(11,43,26,0.25);
      box-shadow:none;
    }
    .btn-secondary:hover{background:var(--brg-light)}
    .btn:disabled{
      opacity:0.55;
      cursor:not-allowed;
      box-shadow:none;
    }

    .muted{color:var(--muted);font-size:14px}
    .sep{height:1px;background:rgba(11,43,26,0.15);border:0;margin:8px 0 0 0}

    .grid{
      flex:1;min-height:0;
      display:grid;
      grid-template-columns:1fr 2fr;
      gap:16px;
    }

    .card{
      border:1px solid rgba(11,43,26,0.12);
      box-shadow:var(--shadow);
      border-radius:18px;
      background:#fff;
      display:flex;
      flex-direction:column;
      min-height:0;
    }
    .card-head{
      padding:14px 16px;
      border-bottom:1px solid rgba(11,43,26,0.10);
      background:linear-gradient(180deg,var(--brg-light),#ffffff);
      border-top-left-radius:18px;
      border-top-right-radius:18px;
    }
    .card-title{margin:0;font-size:16px;font-weight:800;letter-spacing:0.2px;color:var(--brg-dark)}
    .card-body{padding:16px;overflow:auto;min-height:0}

    .field{display:flex;flex-direction:column;gap:8px;margin-bottom:14px}
    label{font-size:13px;font-weight:700;color:var(--brg-dark)}
    input[type="text"],select,textarea{
      border:1px solid var(--border);
      border-radius:12px;
      padding:10px 12px;
      font-size:14px;
      outline:none;
    }
    input[type="text"]:focus,select:focus,textarea:focus{
      box-shadow:0 0 0 3px rgba(20,91,52,0.18);
      border-color:rgba(20,91,52,0.35);
    }
    textarea{min-height:70px;resize:vertical}
    .two-col{display:grid;grid-template-columns:1fr 1fr;gap:12px}

    .right-stack{display:flex;flex-direction:column;gap:16px;min-height:0}
    .bom-card{flex:1;min-height:0}

    .table-wrap{border:1px solid var(--border);border-radius:14px;overflow:auto}

    table{border-collapse:collapse;width:100%;table-layout:fixed;font-size:13px}

    thead th{
      position:sticky;top:0;
      background:var(--brg-light);
      text-align:left;
      padding:10px 12px;
      border-bottom:1px solid rgba(11,43,26,0.14);
      color:var(--brg-dark);
      font-weight:800;
      white-space:nowrap;
      overflow:hidden;
      text-overflow:ellipsis;
    }

    tbody td{
      padding:10px 12px;
      border-bottom:1px solid rgba(11,43,26,0.10);
      color:rgba(11,43,26,0.86);
      white-space:nowrap;
      overflow:hidden;
      text-overflow:ellipsis;
    }

    tbody tr:hover{background:var(--brg-light)}

    .actions-row{display:flex;flex-wrap:wrap;flex-direction: column;gap:8px}
    .divider{height:1px;background:rgba(11,43,26,0.15);margin:14px 0}

    .weighting-head{display:flex;justify-content:space-between;align-items:flex-end;gap:12px;margin-bottom:10px}
    .weighting-title{font-size:14px;font-weight:800}
    .weighting-sub{font-size:12px;color:var(--muted);margin-top:4px}
    .weighting-sub span{color:var(--brg-dark);font-weight:800}
    .hint{font-size:12px;color:var(--muted)}

    .slider-wrap{display:flex;flex-direction:column;gap:8px}
    input[type="range"]{width:100%;accent-color:var(--brg-mid);height:6px}
    .slider-labels{
      display:grid;
      grid-template-columns:1fr 1fr 1fr;
      font-size:11px;
      color:var(--muted);
    }

    .status { margin-top: 10px; }

    .mini-status {
      display:flex;
      gap:10px;
      flex-wrap:wrap;
      margin-top:10px;
      font-size:12px;
      color:var(--muted);
    }
    .badge {
      border:1px solid var(--border);
      background:#fff;
      border-radius:999px;
      padding:4px 10px;
    }
    .badge.ok {
      border-color: rgba(20,91,52,0.35);
      background: rgba(231,243,236,0.9);
      color: var(--brg-dark);
      font-weight:700;
    }

    @media(max-width:1100px){
      .grid{grid-template-columns:1fr}
      .pill{display:none}
    }
  </style>
</head>
<body>
  <div class="app">
    <div class="frame">
      <div class="topbar">
        <label for="fileInput" class="btn btn-primary" id="btnUploadLabel" style="display:inline-block;">
          Upload BOM (Excel)
        </label>

        <div class="muted" id="bomLabel">No BOM loaded</div>

        <div class="pill">UI</div>
        <input id="fileInput" type="file" accept=".xlsx,.xlsm" style="display:none" />
      </div>

      <hr class="sep" />

      <div class="grid">
        <!-- PCB DETAILS -->
        <section class="card">
          <div class="card-head"><h2 class="card-title">PCB Details</h2></div>
          <div class="card-body">
            <div class="field"><label>Board number</label><input type="text" /></div>
            <div class="field"><label>PCB thickness (mm)</label><input type="text" placeholder="e.g. 1.6" /></div>

            <div class="two-col">
              <div class="field"><label>PCB length (mm)</label><input type="text" placeholder="L" /></div>
              <div class="field"><label>PCB width (mm)</label><input type="text" placeholder="W" /></div>
            </div>

            <div class="field">
              <label>Material</label>
              <select>
                <option value="" selected disabled>Select material</option>
                <option>FR-4</option>
                <option>High-Tg FR-4</option>
                <option>Polyimide</option>
                <option>Rogers (RF)</option>
                <option>Aluminium-core</option>
                <option>Other</option>
              </select>
            </div>

            <div class="two-col">
              <div class="field"><label>Layer count</label><input type="text" placeholder="-" /></div>
              <div class="field"><label>Copper weight (oz)</label><input type="text" placeholder="-" /></div>
            </div>

            <div class="two-col">
              <div class="field">
                <label>Finish</label>
                <select>
                  <option value="" selected disabled>Select finish</option>
                  <option>ENIG</option>
                  <option>HASL (Leaded)</option>
                  <option>HASL (Lead-free)</option>
                  <option>OSP</option>
                  <option>Immersion Silver</option>
                  <option>Immersion Tin</option>
                </select>
              </div>
              <div class="field"><label>Tg (if known)</label><input type="text" placeholder="-" /></div>
            </div>

            <div class="field">
              <label>Notes</label>
              <textarea placeholder="Anything relevant: heavy copper, thermal mass, mixed material, specific devices, etc."></textarea>
            </div>

            <div class="divider"></div>

            <div class="actions-row">
            <button class="btn btn-secondary" type="button" id="btnSavePCB">Save PCB details</button>
            <button class="btn btn-secondary" type="button" id="btnLoadPCB">Load PCB details</button>
            <button class="btn btn-secondary" type="button" id="btnClearPCB">Clear PCB details</button>
            </div>

          </div>
        </section>

        <!-- BOM + ACTIONS -->
        <section class="right-stack">
          <div class="card bom-card">
            <div class="card-head"><h2 class="card-title">BOM Preview</h2></div>
            <div class="card-body">
              <div class="table-wrap">
                <table>
                  <thead id="bomHead"></thead>
                  <tbody id="bomRows"></tbody>
                </table>
              </div>

              <div class="muted status" id="statusText">Status: waiting</div>
            </div>
          </div>

          <div class="card">
            <div class="card-head"><h2 class="card-title">Actions</h2></div>
            <div class="card-body">
              <div style="display:flex; gap:10px; flex-wrap:wrap;">
              
                <button class="btn btn-secondary" id="btnBoardAgent" type="button">
                Run board reflow from PCB info
                </button>
                <button class="btn btn-secondary" id="btnComponentAgent" type="button" disabled>
                    Run component reflow agent (MPN-only)
                </button>
                <button class="btn btn-secondary" id="btnViewComponentResults" type="button" disabled>
                    View component results
                </button>
                </div>

                <div class="mini-status">
                <div class="badge" id="badgeBoard">Board: not activated</div>
                <div class="badge" id="badgeComponents">Components: not activated</div>
                </div>

              <div class="divider"></div>

              <div class="weighting">
                <div class="weighting-head">
                  <div>
                    <div class="weighting-title">Weighting</div>
                    <div class="weighting-sub">
                      Board: <span id="boardPct">80%</span> · Components: <span id="compPct">20%</span>
                    </div>
                  </div>
                  <div class="hint">Default 80/20</div>
                </div>

                <div class="slider-wrap">
                  <input id="weightSlider" type="range" min="0" max="100" step="1" value="80" />
                  <div class="slider-labels">
                    <div style="text-align:left;">Components-heavy</div>
                    <div style="text-align:center;">Balanced</div>
                    <div style="text-align:right;">Board-heavy</div>
                  </div>
                </div>
              </div>

              <div class="divider"></div>

              <button class="btn btn-primary" id="btnAggregate" style="width:100%;" type="button" disabled>
                Aggregate reflow profile
              </button>
              <div class="muted status" id="aggStatus">Aggregation: not run</div>
            </div>
          </div>

          <!-- Results modal -->
          <div id="modal" style="display:none; position:fixed; inset:0; background:rgba(0,0,0,0.35);">
            <div style="max-width:1100px; margin:6vh auto; background:#fff; border-radius:18px;
              border:1px solid var(--border); box-shadow: var(--shadow); overflow:hidden;">
              <div style="padding:14px 16px; background:linear-gradient(180deg,var(--brg-light),#fff);
                border-bottom:1px solid rgba(11,43,26,0.10); display:flex; align-items:center; justify-content:space-between;">
                <div style="font-weight:800; color:var(--brg-dark);">Component reflow results (scaffold)</div>
                <button class="btn btn-secondary" id="btnCloseModal" type="button">Close</button>
              </div>
              <div style="padding:16px;">
                <div class="muted" style="margin-bottom:10px;">
                  Fields: Tp, time within 5°C of Tp, TAL, ramp up / cool down (when available).
                  For now these are placeholders until datasheet extraction is implemented.
                </div>
                <div class="table-wrap" style="max-height:60vh;">
                  <table>
                    <thead id="resHead"></thead>
                    <tbody id="resRows"></tbody>
                  </table>
                </div>
              </div>
            </div>
          </div>

        </section>
      </div>
    </div>
  </div>

<script>
document.addEventListener('DOMContentLoaded', () => {
  const bomLabel = document.getElementById('bomLabel');
  const statusText = document.getElementById('statusText');
  const fileInput = document.getElementById('fileInput');
  const bomHead = document.getElementById('bomHead');
  const bomRows = document.getElementById('bomRows');

  const slider = document.getElementById('weightSlider');
  const boardPct = document.getElementById('boardPct');
  const compPct = document.getElementById('compPct');

  const btnComponentAgent = document.getElementById('btnComponentAgent');
  const btnViewComponentResults = document.getElementById('btnViewComponentResults');

  const btnBoardAgent = document.getElementById('btnBoardAgent');
  const pcbStatus = document.getElementById('pcbStatus');

  const badgeComponents = document.getElementById('badgeComponents');
  const badgeBoard = document.getElementById('badgeBoard');

  const btnAggregate = document.getElementById('btnAggregate');
  const aggStatus = document.getElementById('aggStatus');

  const modal = document.getElementById('modal');
  const btnCloseModal = document.getElementById('btnCloseModal');
  const resHead = document.getElementById('resHead');
  const resRows = document.getElementById('resRows');

  function setStatus(msg) {
    if (statusText) statusText.textContent = msg;
  }

  function updatePct(v) {
    const b = Number(v);
    const c = 100 - b;
    if (boardPct) boardPct.textContent = `${b}%`;
    if (compPct) compPct.textContent = `${c}%`;
  }

  function escapeHtml(s) {
    return String(s)
      .replaceAll('&','&amp;')
      .replaceAll('<','&lt;')
      .replaceAll('>','&gt;')
      .replaceAll('"','&quot;')
      .replaceAll("'",'&#039;');
  }

  function renderBomTable(columns, rows) {
    if (!bomHead || !bomRows) return;
    bomHead.innerHTML = `<tr>${columns.map(c => `<th>${escapeHtml(c)}</th>`).join('')}</tr>`;
    bomRows.innerHTML = rows.map(r => `<tr>${columns.map(c => `<td>${escapeHtml(r[c] ?? '')}</td>`).join('')}</tr>`).join('');
  }

  function setBadge(el, ok, text) {
    if (!el) return;
    el.textContent = text;
    el.classList.toggle('ok', !!ok);
  }

  function updateAggregateEnabled(st) {
    const canAgg = !!st.components_ready && !!st.board_ready;
    if (btnAggregate) btnAggregate.disabled = !canAgg;
  }

  // Slider -> persist weighting
  if (slider) {
    slider.addEventListener('input', (e) => {
      updatePct(e.target.value);
      fetch('/api/state', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({ boardWeight: Number(e.target.value) })
      }).catch(() => {});
    });
    updatePct(slider.value);
  }

  // BOM upload
  if (fileInput) {
    fileInput.addEventListener('change', async () => {
      const f = fileInput.files && fileInput.files[0];
      if (!f) return;

      setStatus('Status: uploading BOM...');
      const fd = new FormData();
      fd.append('bom', f, f.name);

      try {
        const res = await fetch('/api/upload_bom', { method: 'POST', body: fd });
        const ct = res.headers.get('content-type') || '';
        const payload = ct.includes('application/json') ? await res.json() : { error: await res.text() };

        if (!res.ok) {
          setStatus('Status: upload failed');
          alert(payload && payload.error ? payload.error : 'Upload failed');
          return;
        }

        if (bomLabel) bomLabel.textContent = payload.label || ('Loaded: ' + f.name);
        renderBomTable(payload.columns || [], payload.rows || []);
        setStatus('Status: BOM loaded');

        if (btnComponentAgent) btnComponentAgent.disabled = (payload.rows || []).length === 0;
        if (btnViewComponentResults) btnViewComponentResults.disabled = true;
      } catch (err) {
        console.error(err);
        setStatus('Status: upload failed (see console)');
      } finally {
        fileInput.value = '';
      }
    });
  }

  // Component agent trigger (MPN-only scaffold)
  if (btnComponentAgent) {
    btnComponentAgent.addEventListener('click', async () => {
      setStatus('Status: running component agent...');
      btnComponentAgent.disabled = true;

      try {
        const res = await fetch('/api/run_component_agent', { method: 'POST' });
        const data = await res.json();
        if (!res.ok) {
          setStatus('Status: component agent failed');
          alert(data && data.error ? data.error : 'Component agent failed');
          btnComponentAgent.disabled = false;
          return;
        }

        setStatus(`Status: component agent complete (${data.unique_mpn_count} unique MPNs)`);
        if (btnViewComponentResults) btnViewComponentResults.disabled = (data.unique_mpn_count === 0);

        const st = await (await fetch('/api/state')).json();
        setBadge(badgeComponents, st.components_ready, st.components_ready ? 'Components: activated' : 'Components: not activated');
        updateAggregateEnabled(st);
      } catch (err) {
        console.error(err);
        setStatus('Status: component agent failed (see console)');
        btnComponentAgent.disabled = false;
      }
    });
  }

  // Board activation (placeholder)
  if (btnBoardAgent) {
    btnBoardAgent.addEventListener('click', async () => {
      if (pcbStatus) pcbStatus.textContent = 'Board: activating...';
      try {
        const res = await fetch('/api/activate_board', { method: 'POST' });
        const data = await res.json();
        if (!res.ok) {
          if (pcbStatus) pcbStatus.textContent = 'Board: activation failed';
          alert(data && data.error ? data.error : 'Board activation failed');
          return;
        }
        if (pcbStatus) pcbStatus.textContent = 'Board: activated (placeholder)';
        const st = await (await fetch('/api/state')).json();
        setBadge(badgeBoard, st.board_ready, st.board_ready ? 'Board: activated' : 'Board: not activated');
        updateAggregateEnabled(st);
      } catch (err) {
        console.error(err);
        if (pcbStatus) pcbStatus.textContent = 'Board: activation failed (see console)';
      }
    });
  }

  // Aggregate (placeholder)
  if (btnAggregate) {
    btnAggregate.addEventListener('click', async () => {
      if (aggStatus) aggStatus.textContent = 'Aggregation: running...';
      try {
        const res = await fetch('/api/aggregate', { method: 'POST' });
        const data = await res.json();
        if (!res.ok) {
          if (aggStatus) aggStatus.textContent = 'Aggregation: failed';
          alert(data && data.error ? data.error : 'Aggregation failed');
          return;
        }
        if (aggStatus) aggStatus.textContent = 'Aggregation: complete (placeholder)';
        alert(data.summary || 'Aggregation complete');
      } catch (err) {
        console.error(err);
        if (aggStatus) aggStatus.textContent = 'Aggregation: failed (see console)';
      }
    });
  }

  // Results modal
  function renderResultsTable(rows) {
    const cols = ["MPN", "Tp (°C)", "Time within 5°C of Tp (s)", "TAL max (s)", "Ramp up (°C/s)", "Cool down (°C/s)", "Status"];
    resHead.innerHTML = `<tr>${cols.map(c => `<th>${escapeHtml(c)}</th>`).join('')}</tr>`;
    resRows.innerHTML = rows.map(r => {
      const td = (v) => `<td>${escapeHtml(v ?? '')}</td>`;
      return `<tr>${
        td(r.mpn) +
        td(r.reflow?.tp_c ?? '') +
        td(r.reflow?.time_within_5c_of_tp_s ?? '') +
        td(r.reflow?.tal_s ?? '') +
        td(r.reflow?.ramp_up_c_per_s ?? '') +
        td(r.reflow?.cool_down_c_per_s ?? '') +
        td(r.status ?? '')
      }</tr>`;
    }).join('');
  }

  if (btnViewComponentResults) {
    btnViewComponentResults.addEventListener('click', async () => {
      try {
        const st = await (await fetch('/api/state')).json();
        renderResultsTable(st.component_reflow || []);
        modal.style.display = 'block';
      } catch (err) {
        console.error(err);
      }
    });
  }
  if (btnCloseModal) btnCloseModal.addEventListener('click', () => modal.style.display = 'none');
  if (modal) modal.addEventListener('click', (e) => { if (e.target === modal) modal.style.display = 'none'; });

  // Initial load
  (async () => {
    try {
      const st = await (await fetch('/api/state')).json();

      if (typeof st.boardWeight === 'number' && slider) {
        slider.value = String(st.boardWeight);
        updatePct(st.boardWeight);
      }

      if (bomLabel && st.bomLabel) bomLabel.textContent = st.bomLabel;

      if (st.bom && Array.isArray(st.bom.columns) && Array.isArray(st.bom.rows)) {
        renderBomTable(st.bom.columns, st.bom.rows);
        if (btnComponentAgent) btnComponentAgent.disabled = (st.bom.rows.length === 0);
      } else {
        renderBomTable(["Item","Stock","Description","Qty","Ref","MPN"], []);
        if (btnComponentAgent) btnComponentAgent.disabled = true;
      }

      setBadge(badgeComponents, st.components_ready, st.components_ready ? 'Components: activated' : 'Components: not activated');
      setBadge(badgeBoard, st.board_ready, st.board_ready ? 'Board: activated' : 'Board: not activated');

      if (pcbStatus) pcbStatus.textContent = st.board_ready ? 'Board: activated (placeholder)' : 'Board: not activated';

      updateAggregateEnabled(st);
      setStatus('Status: waiting');
    } catch (_) {
      renderBomTable(["Item","Stock","Description","Qty","Ref","MPN"], []);
      setStatus('Status: waiting');
    }
  })();
});
</script>
</body>
</html>
"""


def _parse_excel_preview(file_bytes: bytes, filename: str, max_rows: int = 80) -> tuple[str, list[str], list[dict]]:
    lower = filename.lower()
    if not (lower.endswith(".xlsx") or lower.endswith(".xlsm")):
        raise ValueError("Only .xlsx and .xlsm are supported at the moment.")

    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("openpyxl is required. Install with: py -m pip install openpyxl") from e

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header = None
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(cell is not None and str(cell).strip() != "" for cell in row):
            header = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = i
            break

    if header is None:
        raise ValueError("The BOM sheet appears to be empty.")

    while header and header[-1] == "":
        header.pop()

    seen: dict[str, int] = {}
    columns: list[str] = []
    for col in header:
        name = col if col else "Column"
        count = seen.get(name, 0)
        seen[name] = count + 1
        columns.append(name if count == 0 else f"{name}_{count+1}")

    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= header_row_idx:
            continue
        if len(rows) >= max_rows:
            break

        values = list(row or [])
        if len(values) < len(columns):
            values += [None] * (len(columns) - len(values))
        if len(values) > len(columns):
            values = values[: len(columns)]

        if not any(v is not None and str(v).strip() != "" for v in values):
            continue

        d = {c: ("" if v is None else str(v)) for c, v in zip(columns, values)}
        rows.append(d)

    label = f"Loaded: {filename} ({ws.max_row} rows, {len(columns)} cols)"
    return label, columns, rows


def _run_component_agent_mpn_only() -> dict:
    """
    Component agent:
    - Reads BOM from state
    - Extracts unique MPNs from 'MPN' column only
    - For each MPN: scour datasheets online + extract reflow fields
    - If nothing found OR errors -> fill NA 
    """
    with _STATE_LOCK:
        bom = _STATE.get("bom", {})
        cols = bom.get("columns", [])
        rows = bom.get("rows", [])

    if not rows:
        with _STATE_LOCK:
            _STATE["components_ready"] = True
            _STATE["component_reflow"] = []
        return {"unique_mpn_count": 0, "message": "No BOM rows loaded."}

    mpn_col = None
    for c in cols:
        if str(c).strip().lower() == "mpn":
            mpn_col = c
            break
    if mpn_col is None:
        raise ValueError("BOM does not contain an 'MPN' column.")

    mpns: list[str] = []
    for r in rows:
        v = (r.get(mpn_col) or "").strip()
        if not v:
            v = NA
        mpns.append(v)

    unique = sorted(set(mpns), key=lambda s: s.lower())

    results = []
    for mpn in unique:
        # Your rules implemented here
        res = _scour_datasheet_and_extract(mpn)
        results.append(res)

    with _STATE_LOCK:
        _STATE["components_ready"] = True
        _STATE["component_reflow"] = results
        _STATE["aggregation"] = {"status": "not_run", "summary": ""}

    return {"unique_mpn_count": len(unique), "message": "Component agent complete."}


def _aggregate_placeholder() -> str:
    with _STATE_LOCK:
        bw = int(_STATE.get("boardWeight", 80))
        cw = 100 - bw
        comps_ready = bool(_STATE.get("components_ready"))
        board_ready = bool(_STATE.get("board_ready"))
        mpn_count = len(_STATE.get("component_reflow", []) or [])

    if not (comps_ready and board_ready):
        raise ValueError("Both Board and Components must be activated before aggregation.")

    return (
        f"Aggregation placeholder complete.\n\n"
        f"Weighting: Board {bw}% / Components {cw}%\n"
        f"Unique MPNs staged: {mpn_count}\n"
    )


# -----------------------------
# Datasheet scouring + parsing
# -----------------------------

NA = "NA"

# Simple in-memory cache: mpn -> result dict
_SCRAPE_CACHE: dict[str, dict] = {}


def _is_na_mpn(mpn: str) -> bool:
    if mpn is None:
        return True
    v = str(mpn).strip()
    if not v:
        return True
    return v.lower() in {"na", "n/a", "tbd", "unknown", "-"}  # treat these as NA


def _na_reflow_result(mpn: str, status: str = "na") -> dict:
    return {
        "mpn": mpn,
        "status": status,
        "reflow": {
            "tp_c": NA,
            "time_within_5c_of_tp_s": NA,
            "tal_s": NA,
            "ramp_up_c_per_s": NA,
            "cool_down_c_per_s": NA,
        },
        "evidence": {
            "source_url": NA,
            "snippet": NA,
        },
    }


def _http_get(url: str, timeout: int = 12) -> bytes:
    req = urllib.request.Request(
        url,
        headers={
            # basic, non-deceptive UA; many sites block default urllib UA
            "User-Agent": "ReflowProfiler/1.0 (Windows; Python urllib)",
            "Accept": "*/*",
        },
        method="GET",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def _search_candidate_urls(mpn: str, max_results: int = 6) -> list[str]:
    """
    Very lightweight search via DuckDuckGo HTML endpoint.
    Returns a small list of candidate URLs.
    Note: this is best-effort; some networks block DDG.
    """
    q = urllib.parse.quote_plus(f'{mpn} datasheet pdf reflow Tp TAL')
    url = f"https://duckduckgo.com/html/?q={q}"

    html = _http_get(url, timeout=12).decode("utf-8", errors="ignore")

    # DuckDuckGo results commonly appear as: <a class="result__a" href="...">
    links = re.findall(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"', html)
    out: list[str] = []
    for href in links:
        # ddg sometimes returns redirect links; keep as-is (often works)
        if href.startswith("http"):
            out.append(href)
        if len(out) >= max_results:
            break
    return out


def _pick_pdf_link(urls: list[str], mpn: str) -> str | None:
    """
    Prefer direct .pdf links; otherwise return first candidate.
    """
    for u in urls:
        if ".pdf" in u.lower():
            return u
    return urls[0] if urls else None


def _download_pdf_bytes(url: str) -> bytes:
    data = _http_get(url, timeout=18)

    # Some links redirect to HTML; quick sanity check
    head = data[:512].lower()
    if b"%pdf" not in head:
        # Not a real PDF
        raise ValueError("Downloaded content is not a PDF")
    return data


def _pdf_to_text(pdf_bytes: bytes, max_pages: int = 25) -> str:
    """
    Extract text from first N pages.
    """
    reader = PdfReader(io.BytesIO(pdf_bytes))
    texts: list[str] = []
    n = min(len(reader.pages), max_pages)
    for i in range(n):
        try:
            t = reader.pages[i].extract_text() or ""
        except Exception:
            t = ""
        if t:
            texts.append(t)
    return "\n".join(texts)


def _to_seconds(value: float, unit: str) -> float:
    unit = unit.lower()
    if unit in {"s", "sec", "secs", "second", "seconds"}:
        return value
    if unit in {"min", "mins", "minute", "minutes"}:
        return value * 60.0
    return value  # unknown, leave unchanged


def _to_c_per_s(value: float, unit: str) -> float:
    unit = unit.lower().replace(" ", "")
    # °C/s
    if unit in {"c/s", "°c/s", "cpersec", "°cpersec"}:
        return value
    # °C/min
    if unit in {"c/min", "°c/min", "cpermin", "°cpermin"}:
        return value / 60.0
    return value  # unknown


def _first_match_float(pattern: str, text: str) -> tuple[float, tuple[str, ...]] | None:
    m = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
    if not m:
        return None
    try:
        val = float(m.group(1))
    except Exception:
        return None
    groups = tuple(m.groups()[1:])  # remaining groups
    return val, groups


def _extract_reflow_fields_from_text(text: str) -> dict:
    """
    Best-effort regex extraction (RANGE STRINGS):
      - Tp (°C) -> e.g. "245–260 °C" or "260 °C"
      - time within 5°C of Tp (s/min) -> e.g. "20–40 s"
      - TAL (s/min) -> e.g. "60–90 s"
      - ramp up / cool down rate limits -> e.g. "3 °C/s" or "1–3 °C/s" or "60–180 °C/min"

    Returns strings with units, or "NA".
    """
    # Normalise whitespace
    t = re.sub(r"[ \t]+", " ", text)
    t = re.sub(r"\n{2,}", "\n", t)

    def norm_dash(s: str) -> str:
        # unify hyphen variants to en-dash for display consistency
        return s.replace(" - ", "–").replace("-", "–").replace(" to ", "–")

    def find_range_value(patterns: list[str]) -> str:
        """
        Returns the first matched value string, normalised.
        Captures:
          group(1)=first number
          group(2)=optional separator ("-", "–", "to") and second number
          group(3)=unit
        """
        for pat in patterns:
            m = re.search(pat, t, flags=re.IGNORECASE | re.MULTILINE)
            if not m:
                continue

            a = m.group(1)
            b = m.group(2)  # may be None
            unit = (m.group(3) or "").strip()

            if b:
                raw = f"{a}-{b} {unit}".strip()
            else:
                raw = f"{a} {unit}".strip()

            return norm_dash(raw)

        return NA

    # --- Tp (°C) ---
    tp_patterns = [
        # Peak package body temperature / Peak temperature / Tp = 245–260 °C
        r"(?:Peak package body temperature|Peak temperature|Tp)\s*[:=]?\s*(\d{2,3}(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d{2,3}(?:\.\d+)?)?\s*(°?\s*C)",
        # Tp (°C): 260
        r"(?:Tp)\s*\(\s*°?\s*C\s*\)\s*[:=]?\s*(\d{2,3}(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d{2,3}(?:\.\d+)?)?\s*(°?\s*C)?",
    ]
    tp = find_range_value(tp_patterns)
    if tp != NA and "c" not in tp.lower():
        tp = f"{tp} °C"

    # --- Time within 5°C of Tp (tP) ---
    within5_patterns = [
        r"(?:Time within\s*5\s*°?\s*C\s*(?:of)?\s*(?:peak|Tp|T[pP]))\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d+(?:\.\d+)?)?\s*(s|sec|secs|seconds|min|mins|minutes)",
        r"(?:tP|tP\(?\s*within\s*5\s*°?\s*C\)?)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d+(?:\.\d+)?)?\s*(s|sec|secs|seconds|min|mins|minutes)",
    ]
    within5 = find_range_value(within5_patterns)

    # --- TAL ---
    tal_patterns = [
        r"(?:Time above liquidus|TAL)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d+(?:\.\d+)?)?\s*(s|sec|secs|seconds|min|mins|minutes)",
        r"(?:Time above liquidus)\s*\(.*?TAL.*?\)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d+(?:\.\d+)?)?\s*(s|sec|secs|seconds|min|mins|minutes)",
    ]
    tal = find_range_value(tal_patterns)

    # --- Ramp up ---
    ramp_up_patterns = [
        r"(?:Ramp[\- ]?up rate|Heating rate|Ramp rate)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d+(?:\.\d+)?)?\s*(°?\s*C\s*/\s*s|°?\s*C\s*/\s*min)",
        r"(?:Ramp[\- ]?up)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d+(?:\.\d+)?)?\s*(°?\s*C\s*/\s*s|°?\s*C\s*/\s*min)",
    ]
    ramp_up = find_range_value(ramp_up_patterns)
    ramp_up = ramp_up.replace("° C", "°C").replace("C /", "°C/").replace(" / ", "/")

    # --- Cool down ---
    cool_down_patterns = [
        r"(?:Ramp[\- ]?down rate|Cooling rate|Cool(?:ing)? rate)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d+(?:\.\d+)?)?\s*(°?\s*C\s*/\s*s|°?\s*C\s*/\s*min)",
        r"(?:Ramp[\- ]?down|Cool(?:ing)?)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:[–\-]|to)?\s*(\d+(?:\.\d+)?)?\s*(°?\s*C\s*/\s*s|°?\s*C\s*/\s*min)",
    ]
    cool_down = find_range_value(cool_down_patterns)
    cool_down = cool_down.replace("° C", "°C").replace("C /", "°C/").replace(" / ", "/")

    return {
        "tp_c": tp,
        "time_within_5c_of_tp_s": within5,
        "tal_s": tal,
        "ramp_up_c_per_s": ramp_up,
        "cool_down_c_per_s": cool_down,
    }


def _scour_datasheet_and_extract(mpn: str) -> dict:
    """
    Full pipeline (best-effort):
    - search -> pick link -> download pdf -> extract text -> parse fields
    - returns NA if anything fails or no signals found
    """
    mpn_norm = mpn.strip()

    if _is_na_mpn(mpn_norm):
        return _na_reflow_result(mpn_norm, status="mpn_na")

    # cache
    if mpn_norm in _SCRAPE_CACHE:
        return _SCRAPE_CACHE[mpn_norm]

    try:
        urls = _search_candidate_urls(mpn_norm)
        pick = _pick_pdf_link(urls, mpn_norm)
        if not pick:
            res = _na_reflow_result(mpn_norm, status="not_found")
            _SCRAPE_CACHE[mpn_norm] = res
            return res

        pdf_bytes = _download_pdf_bytes(pick)
        text = _pdf_to_text(pdf_bytes, max_pages=25)

        fields = _extract_reflow_fields_from_text(text)

        # If we got no meaningful fields at all, treat as NA
        if all(fields[k] == NA for k in fields):
            res = _na_reflow_result(mpn_norm, status="no_reflow_info")
            res["evidence"]["source_url"] = pick
            _SCRAPE_CACHE[mpn_norm] = res
            return res

        # small snippet for evidence (first occurrence of relevant keywords)
        snippet = NA
        m = re.search(r"(reflow|liquidus|TAL|Tp|peak temperature|ramp)", text, flags=re.IGNORECASE)
        if m:
            start = max(0, m.start() - 120)
            end = min(len(text), m.start() + 220)
            snippet = re.sub(r"\s+", " ", text[start:end]).strip()

        res = {
            "mpn": mpn_norm,
            "status": "ok",
            "reflow": fields,
            "evidence": {
                "source_url": pick,
                "snippet": snippet,
            },
        }
        _SCRAPE_CACHE[mpn_norm] = res
        return res

    except Exception:
        res = _na_reflow_result(mpn_norm, status="error_or_blocked")
        _SCRAPE_CACHE[mpn_norm] = res
        return res

class Handler(BaseHTTPRequestHandler):
    server_version = "ReflowHTTP/1.3"

    def _send_json(self, obj: object, status: int = 200) -> None:
        data = json.dumps(obj).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def _read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        try:
            obj = json.loads(raw.decode("utf-8"))
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}

    def do_GET(self) -> None:
        if self.path in ("/", "/index.html"):
            body = INDEX_HTML.encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(body)
            return

        if self.path == "/api/state":
            with _STATE_LOCK:
                self._send_json(_STATE)
            return

        if self.path == "/api/ping":
            self._send_json({"ok": True, "service": "reflow"})
            return

        self.send_error(HTTPStatus.NOT_FOUND, "Not Found")

    def do_POST(self) -> None:
        if self.path == "/api/state":
            patch = self._read_json()
            with _STATE_LOCK:
                if "boardWeight" in patch:
                    try:
                        bw = int(patch["boardWeight"])
                        _STATE["boardWeight"] = max(0, min(100, bw))
                    except Exception:
                        pass
            self._send_json({"ok": True})
            return

        if self.path == "/api/upload_bom":
            ctype, _ = cgi.parse_header(self.headers.get("Content-Type", ""))
            if ctype != "multipart/form-data":
                self._send_json({"error": "Expected multipart/form-data"}, status=400)
                return

            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )

            if "bom" not in form:
                self._send_json({"error": "No file field named 'bom' provided."}, status=400)
                return

            field = form["bom"]
            if not getattr(field, "file", None):
                self._send_json({"error": "Invalid upload."}, status=400)
                return

            filename = getattr(field, "filename", "") or "bom.xlsx"
            try:
                file_bytes = field.file.read()
                label, columns, rows = _parse_excel_preview(file_bytes, filename, max_rows=80)

                with _STATE_LOCK:
                    _STATE["bomLabel"] = label
                    _STATE["bom"] = {"columns": columns, "rows": rows}
                    _STATE["components_ready"] = False
                    _STATE["component_reflow"] = []
                    _STATE["aggregation"] = {"status": "not_run", "summary": ""}

                self._send_json({"ok": True, "label": label, "columns": columns, "rows": rows})
            except Exception as e:
                self._send_json({"error": str(e)}, status=400)
            return

        if self.path == "/api/run_component_agent":
            try:
                result = _run_component_agent_mpn_only()
                self._send_json({"ok": True, **result})
            except Exception as e:
                self._send_json({"error": str(e)}, status=400)
            return

        if self.path == "/api/activate_board":
            with _STATE_LOCK:
                _STATE["board_ready"] = True
                _STATE["aggregation"] = {"status": "not_run", "summary": ""}
            self._send_json({"ok": True})
            return

        if self.path == "/api/aggregate":
            try:
                summary = _aggregate_placeholder()
                with _STATE_LOCK:
                    _STATE["aggregation"] = {"status": "complete", "summary": summary}
                self._send_json({"ok": True, "summary": summary})
            except Exception as e:
                self._send_json({"error": str(e)}, status=400)
            return

        self.send_error(HTTPStatus.NOT_FOUND, "Not Found")

    def log_message(self, fmt: str, *args) -> None:
        return


def _open_browser_later(url: str) -> None:
    time.sleep(0.25)
    try:
        webbrowser.open(url, new=1)
    except Exception:
        pass


def main() -> None:
    url = f"http://{HOST}:{PORT}/"
    httpd = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"Serving on {url}")
    print("Press Ctrl+C to stop.")
    threading.Thread(target=_open_browser_later, args=(url,), daemon=True).start()
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        httpd.server_close()


if __name__ == "__main__":
    main()